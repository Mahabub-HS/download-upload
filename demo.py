import os
import youtube_dl
import openpyxl
from pydrive.drive import GoogleDrive
from pydrive.auth import GoogleAuth

def upload():
    #for uploading into gdrive
    print("---------------Upload into gdrive---------------")
    upload = str(input("Do you upload your video into Google Drive: "))
    if (upload == "yes"):
        gauth = GoogleAuth()
        gauth.LocalWebserverAuth()
        drive = GoogleDrive(gauth)
        path = r"F:\Python\Download & Uploader App\app"
        for x in os.listdir(path):
            f = drive.CreateFile({'title': x,
                                  'parents': [{u'id': '1UGL8t5d_vZlKRCTYQUjOFf7EBcP2Wz71'}]})
            f.SetContentFile(os.path.join(path, x))
            f.Upload()
    else:
        print("Thank you...!")






y = {}
path = "C:\\Users\\MAHABUB\\PycharmProjects\\pythonProject\\Dvideo"
os.chdir(path)

url = input("Url: ")
with youtube_dl.YoutubeDL(y) as u:
    print("Downloading....." + url)
    u.download([url])
    print("Done")
    upload()


# wbk = 'test.xlsx'
# wb = openpyxl.load_workbook(wbk)
# sh1 = wb['Sheet1']
#
#
#
# def write_ok():
#     r = sh1.max_row
#     c = sh1.max_column
#     for i in range(2, r):
#         for j in range(2,3):
#             sh1.cell(row=i,column=j,value='Ok')
#             print(sh1.cell(i,j).value)
#             wb.save(wbk)
# write_ok()


# wbk = 'test.xlsx'
# wb = openpyxl.load_workbook(wbk)
# sh1 = wb['Sheet1']
# r = sh1.max_row
# for i in range(2,r):
#     for j in range(1,2):
#         link = (sh1.cell(i,j).value)
#         down()
#         for i in range(2, r):
#             for j in range(2,3):
#                 sh1.cell(row=i,column=j,value='Ok')
#                 # print(sh1.cell(i,j).value)
#                 wb.save(wbk)
#         print("Successfully ok")
# upload()



